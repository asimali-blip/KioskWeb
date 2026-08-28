#!/usr/bin/env python3
"""
Wavetec Kiosk Bulk Tool — Web Dashboard
========================================
Runs on this Windows machine (same one that can reach the kiosks over
SSH). Provides a browser-based control panel:

  - Login system (multi-user, roles: admin / operator)
  - Dashboard with summary cards from the last completed run
  - Device list management (add/edit/delete/import CSV)
  - Password list management
  - Settings (ports, timeouts, remote paths, search/replace text, retry, backup)
  - Run control: pick jobs, start a run, watch live logs stream in the browser
  - Results viewer (searchable table, same data as results.csv)
  - User management (admin only)

First run: creates a default admin user (username: admin / password: admin)
and forces a password change on first login.
"""

import os
import json
import time
import threading
import queue
import functools
import secrets

from flask import (
    Flask, render_template, request, redirect, url_for, session,
    flash, Response, jsonify, send_file
)
from werkzeug.security import generate_password_hash, check_password_hash
from werkzeug.utils import secure_filename

import kiosk_core as core

BASE_DIR = core.BASE_DIR
USERS_FILE = os.path.join(BASE_DIR, "users.json")

app = Flask(__name__,
           template_folder=os.path.join(core.BASE_DIR, "templates"),
           static_folder=os.path.join(core.BASE_DIR, "static"))
app.secret_key = os.environ.get("KIOSK_WEB_SECRET") or secrets.token_hex(32)

# SECURITY: harden the session cookie. Flask's own defaults leave
# SESSION_COOKIE_SECURE off (fine over plain HTTP, but means the cookie
# would still be sent in the clear if someone points a browser at this
# app over HTTP even after HTTPS is set up) and don't set SameSite to a
# strict-enough value for a login-protected internal tool.
app.config.update(
    SESSION_COOKIE_HTTPONLY=True,
    SESSION_COOKIE_SAMESITE="Lax",
    # Auto-detect HTTPS: only mark the cookie Secure when the app is
    # actually being served over TLS (directly, or behind Caddy/IIS which
    # both terminate TLS and this app still runs on plain HTTP locally).
    # Set KIOSK_WEB_FORCE_SECURE_COOKIE=1 once you've put Caddy/IIS HTTPS
    # in front of this app to turn it on.
    SESSION_COOKIE_SECURE=os.environ.get("KIOSK_WEB_FORCE_SECURE_COOKIE") == "1",
    PERMANENT_SESSION_LIFETIME=8 * 60 * 60,  # 8 hours
)

# SECURITY: cap request/upload size so a malicious or mistaken huge upload
# (patch file, APK, config zip, device CSV) can't fill the disk or exhaust
# memory. 500 MB comfortably covers legitimate patch/APK files; raise via
# env var if a real deployment needs bigger firmware bundles.
app.config["MAX_CONTENT_LENGTH"] = int(os.environ.get("KIOSK_WEB_MAX_UPLOAD_MB", "500")) * 1024 * 1024

# ==============================================================================
# Basic login throttling
# SECURITY: there was previously NO limit on login attempts — an attacker
# on the network could brute-force the admin/operator password with an
# unlimited number of guesses. This is a simple in-memory lockout; it
# resets on restart and isn't shared across multiple worker processes,
# but for this single-process waitress deployment it stops naive brute
# forcing without adding a dependency.
# ==============================================================================

_LOGIN_ATTEMPTS = {}         # ip -> [timestamps of recent failures]
_LOGIN_LOCK = threading.Lock()
_LOGIN_MAX_FAILURES = 5
_LOGIN_WINDOW_SECONDS = 300      # 5 minutes
_LOGIN_LOCKOUT_SECONDS = 300     # 5 minutes


def _client_ip():
    return request.remote_addr or "unknown"


def _login_is_locked_out(ip):
    with _LOGIN_LOCK:
        attempts = [t for t in _LOGIN_ATTEMPTS.get(ip, []) if time.time() - t < _LOGIN_WINDOW_SECONDS]
        _LOGIN_ATTEMPTS[ip] = attempts
        return len(attempts) >= _LOGIN_MAX_FAILURES


def _login_record_failure(ip):
    with _LOGIN_LOCK:
        _LOGIN_ATTEMPTS.setdefault(ip, []).append(time.time())


def _login_clear(ip):
    with _LOGIN_LOCK:
        _LOGIN_ATTEMPTS.pop(ip, None)

# ==============================================================================
# Users
# ==============================================================================

DEFAULT_ADMIN_PASSWORD = "admin"

def load_users():
    if not os.path.exists(USERS_FILE):
        users = {
            "admin": {
                "password_hash": generate_password_hash(DEFAULT_ADMIN_PASSWORD),
                "role": "admin",
                "must_change_password": True,
            }
        }
        save_users(users)
        return users
    with open(USERS_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_users(users):
    with open(USERS_FILE, "w", encoding="utf-8") as f:
        json.dump(users, f, indent=2)


def login_required(view):
    @functools.wraps(view)
    def wrapped(*args, **kwargs):
        if "username" not in session:
            return redirect(url_for("login", next=request.path))
        return view(*args, **kwargs)
    return wrapped


def admin_required(view):
    @functools.wraps(view)
    def wrapped(*args, **kwargs):
        if "username" not in session:
            return redirect(url_for("login", next=request.path))
        users = load_users()
        if users.get(session["username"], {}).get("role") != "admin":
            flash("Admin access required for that page.", "error")
            return redirect(url_for("dashboard"))
        return view(*args, **kwargs)
    return wrapped


def local_file_status(path):
    if os.path.exists(path):
        size = os.path.getsize(path)
        mtime = time.strftime("%Y-%m-%d %H:%M", time.localtime(os.path.getmtime(path)))
        return {"exists": True, "size": size, "mtime": mtime}
    return {"exists": False, "size": 0, "mtime": ""}


@app.context_processor
def inject_user():
    username = session.get("username")
    role = None
    if username:
        role = load_users().get(username, {}).get("role")
    return {"current_user": username, "current_role": role, "run_state": RUN_STATE}


# ==============================================================================
# Run state — one batch runs at a time, shared across requests
# ==============================================================================

class RunState:
    def __init__(self):
        self.lock = threading.Lock()
        self.running = False
        self.thread = None
        self.stop_event = threading.Event()
        self.log_lines = []
        self.log_queue_subscribers = []   # list of queue.Queue, one per connected browser tab
        self.summary = None
        self.started_by = None
        self.started_at = None
        self.last_selection = {"run_ppt": False, "run_shot": False, "run_fr": False,
                                "run_patch": False, "run_chain": False, "run_cfu": False, "resume": True,
                                "reboot_patch": True,
                                "device_types": [],
                                "site_filter": []}

    def log(self, msg):
        line = f"{time.strftime('%H:%M:%S')} {msg}"
        with self.lock:
            self.log_lines.append(line)
            if len(self.log_lines) > 5000:
                self.log_lines = self.log_lines[-5000:]
            for q in self.log_queue_subscribers:
                q.put(line)

    def progress(self, done, total, label):
        pct = int(done / total * 100) if total else 100
        self.log(f"PROGRESS {label} {done}/{total} ({pct}%)")

    def subscribe(self):
        q = queue.Queue()
        with self.lock:
            self.log_queue_subscribers.append(q)
        return q

    def unsubscribe(self, q):
        with self.lock:
            if q in self.log_queue_subscribers:
                self.log_queue_subscribers.remove(q)

    def start(self, run_ppt, run_shot, run_fr, run_patch, run_chain, resume, device_types,
              reboot_patch, username, site_filter=None, only_ips=None, run_cfu=False):
        with self.lock:
            if self.running:
                return False
            self.running = True
            self.log_lines = []
            self.summary = None
            self.stop_event = threading.Event()
            self.started_by = username
            self.started_at = time.strftime("%Y-%m-%d %H:%M:%S")

        def _worker():
            try:
                runner = core.BatchRunner(
                    run_ppt, run_shot, run_fr, run_patch, run_chain,
                    log_fn=self.log, progress_fn=self.progress,
                    stop_event=self.stop_event, resume=resume,
                    device_types=device_types, reboot_patch=reboot_patch,
                    site_filter=site_filter, only_ips=only_ips, run_cfu=run_cfu,
                )
                summary = runner.run()
                with self.lock:
                    self.summary = summary
            except Exception as e:
                self.log(f"[FATAL ERROR] {e}")
            finally:
                with self.lock:
                    self.running = False

        self.thread = threading.Thread(target=_worker, daemon=True)
        self.thread.start()
        return True

    def request_stop(self):
        self.stop_event.set()
        self.log("Stop requested — finishing in-flight devices, no new ones will start...")


RUN_STATE = RunState()

# SECURITY: patch/APK uploads are saved directly into BASE_DIR under
# whatever (sanitized) name the file has. secure_filename() strips path
# separators so this isn't traversal, but a file uploaded as e.g.
# "users.json", "settings.json" or "app.py" would silently overwrite that
# real app file. Block writing over anything the app itself depends on.
_PROTECTED_FILENAMES = {
    "users.json", "settings.json", "cfu_settings.json", "patch_chains.json",
    "schedule.json", "kiosk_devices.csv", "passwords.txt", "results.csv",
    "results_inprogress.csv", "app.py", "kiosk_core.py", "requirements.txt",
    "web.config", "caddyfile", "kiosk_web.spec",
}


# ==============================================================================
# CSRF protection
# SECURITY: none of the POST forms in this app carried a CSRF token. Since
# auth is a plain session cookie, any page the logged-in admin/operator
# visits (or any auto-submitting HTML an attacker gets them to open) could
# silently submit a POST to this app — e.g. add an admin user, delete
# devices, change Settings/reboot commands, or (before the fix above)
# delete arbitrary folders. This adds a standard synchronizer-token check:
# every POST must include a "csrf_token" form field matching the one tied
# to the user's session.
#
# NOTE: templates/*.html weren't included in what was reviewed here, so
# each <form method="post"> needs one line added inside it:
#     <input type="hidden" name="csrf_token" value="{{ csrf_token() }}">
# (the SSE /run/stream and JSON /run/status GET endpoints are unaffected —
# CSRF only matters for state-changing POST/DELETE/PUT requests.)
# ==============================================================================

def _get_csrf_token():
    if "csrf_token" not in session:
        session["csrf_token"] = secrets.token_hex(32)
    return session["csrf_token"]


@app.context_processor
def inject_csrf_token():
    return {"csrf_token": _get_csrf_token}


@app.before_request
def _csrf_protect():
    if request.method == "POST":
        # Login itself has no prior session to anchor a token to, and
        # brute-force there is already mitigated separately above.
        if request.endpoint == "login":
            return None
        sent = request.form.get("csrf_token", "")
        expected = session.get("csrf_token", "")
        if not sent or not expected or not secrets.compare_digest(sent, expected):
            flash("Your session expired or the form was out of date — please try again.", "error")
            return redirect(request.referrer or url_for("dashboard"))
    return None


# ==============================================================================
# Auth routes
# ==============================================================================

@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        ip = _client_ip()
        if _login_is_locked_out(ip):
            flash("Too many failed login attempts. Try again in a few minutes.", "error")
            return render_template("login.html")

        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        users = load_users()
        user = users.get(username)
        if user and check_password_hash(user["password_hash"], password):
            _login_clear(ip)
            session.clear()
            session["username"] = username
            session.permanent = True
            if user.get("must_change_password"):
                return redirect(url_for("change_password"))
            nxt = request.args.get("next") or url_for("dashboard")
            return redirect(nxt)
        _login_record_failure(ip)
        flash("Invalid username or password.", "error")
    return render_template("login.html")


@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))


@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    if request.method == "POST":
        new_pw = request.form.get("new_password", "")
        confirm = request.form.get("confirm_password", "")
        if len(new_pw) < 4:
            flash("Password must be at least 4 characters.", "error")
        elif new_pw != confirm:
            flash("Passwords do not match.", "error")
        else:
            users = load_users()
            users[session["username"]]["password_hash"] = generate_password_hash(new_pw)
            users[session["username"]]["must_change_password"] = False
            save_users(users)
            flash("Password updated.", "success")
            return redirect(url_for("dashboard"))
    return render_template("change_password.html")


# ==============================================================================
# Dashboard
# ==============================================================================

@app.route("/")
@login_required
def dashboard():
    rows = core.load_latest_results()
    total = len(rows)

    def count(field, val):
        return sum(1 for r in rows if r.get(field) == val)

    stats = {
        "total": total,
        "ppt_ok": count("pptconfig_status", "SUCCESS"),
        "ppt_fail": count("pptconfig_status", "FAILURE"),
        "ppt_skip": count("pptconfig_status", "SKIPPED"),
        "shot_ok": count("screenshot_status", "SUCCESS"),
        "shot_fail": count("screenshot_status", "FAILURE"),
        "shot_skip": count("screenshot_status", "SKIPPED"),
        "fr_ok": count("filereplace_status", "SUCCESS"),
        "fr_fail": count("filereplace_status", "FAILURE"),
        "fr_skip": count("filereplace_status", "SKIPPED"),
        "patch_ok": count("patch_status", "SUCCESS"),
        "patch_fail": count("patch_status", "FAILURE"),
        "patch_skip": count("patch_status", "SKIPPED"),
    }
    device_count = len(core.load_devices())
    password_count = len(core.load_passwords())
    return render_template("dashboard.html", stats=stats, device_count=device_count,
                            password_count=password_count)


# ==============================================================================
# Devices
# ==============================================================================

import re
_IPV4_RE = re.compile(r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$")

def looks_like_ip(value):
    """True if value looks like a plausible IPv4 address (not a hostname/typo)."""
    if not _IPV4_RE.match(value or ""):
        return False
    return all(0 <= int(p) <= 255 for p in value.split("."))


@app.route("/devices", methods=["GET", "POST"])
@login_required
def devices():
    if request.method == "POST":
        action = request.form.get("action")
        devs = core.load_devices()
        if action == "add":
            name = request.form.get("device_name", "").strip()
            ip = request.form.get("ip", "").strip()
            site = request.form.get("site", "").strip()
            device_type = request.form.get("device_type", "").strip()
            if device_type not in core.DEVICE_TYPES:
                device_type = core.DEFAULT_DEVICE_TYPE
            if ip:
                devs.append({"device_name": name or ip, "ip": ip, "device_type": device_type, "site": site})
                core.save_devices(devs)
                if looks_like_ip(ip):
                    flash(f"Added {name or ip}.", "success")
                else:
                    flash(f"Added {name or ip} — but \"{ip}\" doesn't look like a normal IP "
                          f"address (e.g. 10.101.13.196). If it's meant to be an IP, double-check "
                          f"the device_name and IP fields weren't swapped. If it's a hostname, "
                          f"make sure this machine can resolve it (DNS or hosts file).", "error")
        elif action == "delete":
            ip_to_delete = request.form.get("ip")
            devs = [d for d in devs if d["ip"] != ip_to_delete]
            core.save_devices(devs)
            flash("Device removed.", "success")
        elif action == "import_csv":
            file = request.files.get("csv_file")
            if file and file.filename:
                file.save(core.DEVICES_FILE)
                imported = core.load_devices()
                bad_rows = [d for d in imported if not looks_like_ip(d["ip"])]
                if bad_rows:
                    preview = ", ".join(f'{d["device_name"]}→"{d["ip"]}"' for d in bad_rows[:5])
                    more = f" (+{len(bad_rows) - 5} more)" if len(bad_rows) > 5 else ""
                    flash(f"Imported {len(imported)} device(s), but {len(bad_rows)} row(s) don't "
                          f"look like valid IP addresses — check for swapped columns: "
                          f"{preview}{more}", "error")
                else:
                    flash(f"Imported {len(imported)} device(s).", "success")
        return redirect(url_for("devices"))

    devs = core.load_devices()
    flagged_ips = {d["ip"] for d in devs if not looks_like_ip(d["ip"])}
    return render_template("devices.html", devices=devs, flagged_ips=flagged_ips,
                           device_types=core.DEVICE_TYPES, sites=core.distinct_sites())


@app.route("/devices/bulk", methods=["POST"])
@login_required
def devices_bulk():
    action = request.form.get("action")
    selected_ips = request.form.getlist("selected_ips")

    if not selected_ips:
        flash("No devices selected.", "error")
        return redirect(url_for("devices"))

    if action == "delete":
        devs = core.load_devices()
        remaining = [d for d in devs if d["ip"] not in selected_ips]
        removed_count = len(devs) - len(remaining)
        core.save_devices(remaining)
        flash(f"Removed {removed_count} device(s).", "success")
        return redirect(url_for("devices"))

    if action == "run":
        return redirect(url_for("run_page", only_ip=selected_ips))

    flash("Unknown bulk action.", "error")
    return redirect(url_for("devices"))


# ==============================================================================
# Passwords
# ==============================================================================

@app.route("/passwords", methods=["GET", "POST"])
@login_required
def passwords():
    if request.method == "POST":
        raw = request.form.get("passwords", "")
        pw_list = [p.strip() for p in raw.splitlines() if p.strip()]
        core.save_passwords(pw_list)
        flash("Password list updated.", "success")
        return redirect(url_for("passwords"))

    pw_list = core.load_passwords()
    return render_template("passwords.html", passwords=pw_list)


# ==============================================================================
# Settings (admin only)
# ==============================================================================

@app.route("/settings", methods=["GET", "POST"])
@admin_required
def settings_page():
    if request.method == "POST":
        s = core.load_settings()
        s["username"] = request.form.get("username", s["username"]).strip()
        s["ports"] = [int(p.strip()) for p in request.form.get("ports", "").split(",") if p.strip().isdigit()]
        s["remote_pptconfig_file"] = request.form.get("remote_pptconfig_file", "").strip()
        s["search_text"] = request.form.get("search_text", "").strip()
        s["replace_text"] = request.form.get("replace_text", "").strip()
        s["remote_screen_path"] = request.form.get("remote_screen_path", "").strip()
        s["remote_slide_file"] = request.form.get("remote_slide_file", "").strip()
        s["remote_patch_file"] = request.form.get("remote_patch_file", "").strip()
        s["patch_filename"] = request.form.get("patch_filename", "patch.zip").strip() or "patch.zip"
        s["reboot_command"] = request.form.get("reboot_command", "reboot").strip()
        s["reboot_after_patch"] = request.form.get("reboot_after_patch") == "on"
        for key in ("threads", "chain_threads", "ssh_timeout", "scp_timeout", "connect_timeout", "retry_passes"):
            try:
                s[key] = int(request.form.get(key, s[key]))
            except ValueError:
                pass
        s["enable_backup"] = request.form.get("enable_backup") == "on"
        core.save_settings(s)
        flash("Settings saved.", "success")
        return redirect(url_for("settings_page"))

    s = core.load_settings()
    patch_status = local_file_status(core.get_local_patch_path(s))
    slide_status = local_file_status(core.LOCAL_SLIDE_SOURCE)
    return render_template("settings.html", s=s, patch_status=patch_status, slide_status=slide_status)


@app.route("/settings/upload_patch", methods=["POST"])
@admin_required
def upload_patch():
    file = request.files.get("patch_file")
    if not file or not file.filename:
        flash("No patch file selected.", "error")
        return redirect(url_for("settings_page"))
    filename = secure_filename(file.filename)
    if not filename or filename.lower() in _PROTECTED_FILENAMES:
        flash("That file name isn't valid.", "error")
        return redirect(url_for("settings_page"))
    dest = os.path.join(core.BASE_DIR, filename)
    file.save(dest)
    s = core.load_settings()
    s["patch_filename"] = filename
    core.save_settings(s)
    flash(f"Uploaded '{filename}' ({os.path.getsize(dest):,} bytes) and set it as the patch file.", "success")
    return redirect(url_for("settings_page"))


@app.route("/settings/upload_slide", methods=["POST"])
@admin_required
def upload_slide():
    file = request.files.get("slide_file")
    if not file or not file.filename:
        flash("No slide image selected.", "error")
        return redirect(url_for("settings_page"))
    dest = core.LOCAL_SLIDE_SOURCE
    file.save(dest)
    flash(f"Uploaded slide image ({os.path.getsize(dest):,} bytes).", "success")
    return redirect(url_for("settings_page"))


# ==============================================================================
# Run control
# ==============================================================================

@app.route("/run")
@login_required
def run_page():
    only_ips = [ip for ip in request.args.getlist("only_ip") if ip.strip()]
    only_devices = []
    if only_ips:
        all_devs = {d["ip"]: d for d in core.load_devices()}
        only_devices = [all_devs.get(ip, {"device_name": ip, "ip": ip, "device_type": "", "site": ""})
                        for ip in only_ips]
    return render_template("run.html", device_types=core.DEVICE_TYPES, sites=core.distinct_sites(),
                           only_ips=only_ips, only_devices=only_devices)


@app.route("/run/start", methods=["POST"])
@login_required
def run_start():
    run_ppt = request.form.get("run_ppt") == "on"
    run_shot = request.form.get("run_shot") == "on"
    run_fr = request.form.get("run_fr") == "on"
    run_patch = request.form.get("run_patch") == "on"
    run_chain = request.form.get("run_chain") == "on"
    run_cfu = request.form.get("run_cfu") == "on"
    resume = request.form.get("resume") == "on"
    reboot_patch = request.form.get("reboot_patch") == "on"
    device_types = request.form.getlist("device_types")
    if not device_types:
        device_types = list(core.DEVICE_TYPES)  # none checked = treat as "all"
    site_filter = request.form.getlist("site_filter")  # empty = all sites
    only_ip_list = [ip for ip in request.form.getlist("only_ip") if ip.strip()]
    only_ips = set(only_ip_list) if only_ip_list else None

    RUN_STATE.last_selection = {"run_ppt": run_ppt, "run_shot": run_shot, "run_fr": run_fr,
                                 "run_patch": run_patch, "run_chain": run_chain, "run_cfu": run_cfu,
                                 "resume": resume, "reboot_patch": reboot_patch,
                                 "device_types": device_types, "site_filter": site_filter}

    if not (run_ppt or run_shot or run_fr or run_patch or run_chain or run_cfu):
        flash("Select at least one job.", "error")
        return redirect(url_for("run_page"))

    ok = RUN_STATE.start(run_ppt, run_shot, run_fr, run_patch, run_chain, resume,
                         device_types, reboot_patch, session["username"],
                         site_filter=site_filter, only_ips=only_ips, run_cfu=run_cfu)
    if not ok:
        flash("A run is already in progress.", "error")
    return redirect(url_for("run_page"))


@app.route("/run/stop", methods=["POST"])
@login_required
def run_stop():
    RUN_STATE.request_stop()
    return redirect(url_for("run_page"))


@app.route("/run/status")
@login_required
def run_status():
    with RUN_STATE.lock:
        return jsonify({
            "running": RUN_STATE.running,
            "started_by": RUN_STATE.started_by,
            "started_at": RUN_STATE.started_at,
            "summary": RUN_STATE.summary,
        })


@app.route("/run/stream")
@login_required
def run_stream():
    def gen():
        q = RUN_STATE.subscribe()
        try:
            with RUN_STATE.lock:
                backlog = list(RUN_STATE.log_lines[-200:])
            for line in backlog:
                yield f"data: {line}\n\n"
            while True:
                try:
                    line = q.get(timeout=15)
                    yield f"data: {line}\n\n"
                except queue.Empty:
                    yield ": keepalive\n\n"
        finally:
            RUN_STATE.unsubscribe(q)

    return Response(gen(), mimetype="text/event-stream")


# ==============================================================================
# Results
# ==============================================================================

@app.route("/results")
@login_required
def results():
    rows = core.load_latest_results()
    return render_template("results.html", rows=rows)


@app.route("/results/download")
@login_required
def results_download():
    if not os.path.exists(core.RESULTS_CSV):
        flash("No results yet.", "error")
        return redirect(url_for("results"))
    return send_file(core.RESULTS_CSV, as_attachment=True, download_name="results.csv")


@app.route("/results/download_excel")
@login_required
def results_download_excel():
    rows = core.load_latest_results()
    if not rows:
        flash("No results yet.", "error")
        return redirect(url_for("results"))

    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"

    headers = ["Device", "IP", "Type", "Site", "Android", "Port",
               "PPT Config", "Screenshot", "Slide", "Patch", "Chain", "Time"]
    ws.append(headers)
    header_fill = PatternFill(start_color="1F2937", end_color="1F2937", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    status_fill = {
        "SUCCESS": PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid"),
        "FAILURE": PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid"),
        "SKIPPED": PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid"),
    }
    status_cols = [7, 8, 9, 10, 11]  # PPT/Screenshot/Slide/Patch/Chain columns (1-indexed)

    for row in rows:
        ws.append([
            row.get("device_name", ""), row.get("ip", ""), row.get("device_type", ""),
            row.get("site", ""), row.get("android_version", ""), row.get("port", ""),
            row.get("pptconfig_status", ""), row.get("screenshot_status", ""),
            row.get("filereplace_status", ""), row.get("patch_status", ""),
            row.get("chain_status", ""), row.get("timestamp", ""),
        ])
        r = ws.max_row
        for col in status_cols:
            val = ws.cell(row=r, column=col).value
            fill = status_fill.get(val)
            if fill:
                ws.cell(row=r, column=col).fill = fill

    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 16
    ws.freeze_panes = "A2"

    out_path = os.path.join(core.BASE_DIR, "results_export.xlsx")
    wb.save(out_path)
    return send_file(out_path, as_attachment=True, download_name="results.xlsx")


@app.route("/gallery")
@login_required
def gallery():
    rows = [r for r in core.load_latest_results() if r.get("screenshot_status") == "SUCCESS" and r.get("screenshot_file")]
    for r in rows:
        r["screenshot_basename"] = os.path.basename(r["screenshot_file"])
    return render_template("gallery.html", rows=rows)


@app.route("/screenshots/<path:filename>")
@login_required
def serve_screenshot(filename):
    # SECURITY: filename comes straight from the URL. Without this check,
    # a value like "../../../../Windows/win.ini" (or an absolute path /
    # drive letter) would let ANY logged-in user (including "operator"
    # role) read arbitrary files off the server via path traversal.
    # werkzeug.utils.safe_join resolves the path and returns None if it
    # would escape LOCAL_DIR.
    from werkzeug.utils import safe_join
    safe_path = safe_join(core.LOCAL_DIR, filename)
    if safe_path is None or not os.path.isfile(safe_path):
        return ("Not found", 404)
    return send_file(safe_path)


# ==============================================================================
# Users (admin only)
# ==============================================================================

@app.route("/users", methods=["GET", "POST"])
@admin_required
def users_page():
    users = load_users()
    if request.method == "POST":
        action = request.form.get("action")
        if action == "add":
            username = request.form.get("username", "").strip()
            password = request.form.get("password", "")
            role = request.form.get("role", "operator")
            if not username or not password:
                flash("Username and password required.", "error")
            elif username in users:
                flash("That username already exists.", "error")
            else:
                users[username] = {
                    "password_hash": generate_password_hash(password),
                    "role": role,
                    "must_change_password": True,
                }
                save_users(users)
                flash(f"User '{username}' created.", "success")
        elif action == "delete":
            username = request.form.get("username")
            if username == session["username"]:
                flash("You cannot delete your own account while logged in.", "error")
            elif username in users:
                del users[username]
                save_users(users)
                flash("User removed.", "success")
        return redirect(url_for("users_page"))

    return render_template("users.html", users=users)


# ==============================================================================
# Patch Chains (admin only) — configure per-controller-type sequential
# patch chains for Job 5 (Controller Patch Chain)
# ==============================================================================

@app.route("/patch-chains")
@admin_required
def patch_chains_page():
    chains = core.load_patch_chains()
    available_versions = {ctype: core.scan_available_versions(ctype) for ctype in chains}
    return render_template("patch_chains.html", chains=chains, available_versions=available_versions)


@app.route("/patch-chains/save_config", methods=["POST"])
@admin_required
def patch_chains_save_config():
    ctype = request.form.get("controller_type")
    chains = core.load_patch_chains()
    if ctype not in chains:
        flash("Unknown controller type.", "error")
        return redirect(url_for("patch_chains_page"))

    cfg = chains[ctype]
    cfg["kernel_keyword"] = request.form.get("kernel_keyword", "").strip()
    cfg["version_file"] = request.form.get("version_file", "/eq/version.txt").strip()
    cfg["remote_work_dir"] = request.form.get("remote_work_dir", "/eq").strip()
    cfg["readonly"] = request.form.get("readonly") == "on"
    cfg["remount_rw_command"] = request.form.get("remount_rw_command", "").strip()
    cfg["remount_ro_command"] = request.form.get("remount_ro_command", "").strip()
    try:
        cfg["reboot_wait_seconds"] = int(request.form.get("reboot_wait_seconds", 300))
    except ValueError:
        pass

    core.save_patch_chains(chains)
    flash(f"Saved configuration for {ctype}.", "success")
    return redirect(url_for("patch_chains_page"))


@app.route("/patch-chains/add_step", methods=["POST"])
@admin_required
def patch_chains_add_step():
    """
    Uploads a patch file into ControllerPatches/<Type>/<version>/upgrade.tar.gz.
    No separate registration is needed — do_controller_patch_chain scans this
    folder directly, so versions dropped in here manually (bulk copy, USB,
    network share) work exactly the same as versions added through this form.
    """
    ctype = request.form.get("controller_type")
    version = request.form.get("version", "").strip()
    file = request.files.get("patch_file")

    if not version:
        flash("Version is required (this becomes the folder name).", "error")
        return redirect(url_for("patch_chains_page"))
    if not core.parse_version_tuple(version):
        flash(f"'{version}' doesn't look like a version number (e.g. 18.8.9).", "error")
        return redirect(url_for("patch_chains_page"))
    # SECURITY: parse_version_tuple() uses re.search(), so it only checks
    # that a version-looking substring exists somewhere in the string —
    # "../../../../evil/18.8.9" passes that check while still being a
    # traversal payload for the os.path.join() below. Reject anything
    # that isn't a bare, safe folder-name token before it touches disk.
    if os.path.basename(version) != version or version in (".", "..") or "/" in version or "\\" in version:
        flash(f"'{version}' isn't a valid version folder name.", "error")
        return redirect(url_for("patch_chains_page"))
    if not file or not file.filename:
        flash("A patch file is required.", "error")
        return redirect(url_for("patch_chains_page"))

    chains = core.load_patch_chains()
    if ctype not in chains:
        flash("Unknown controller type.", "error")
        return redirect(url_for("patch_chains_page"))

    if version in core.scan_available_versions(ctype):
        flash(f"Version {version} already exists for {ctype} — remove it first to replace.", "error")
        return redirect(url_for("patch_chains_page"))

    base = os.path.realpath(core.CONTROLLER_PATCHES_DIR)
    version_dir = os.path.realpath(os.path.join(base, ctype.replace(" ", "_"), version))
    if not version_dir.startswith(base + os.sep):
        flash("Invalid version path.", "error")
        return redirect(url_for("patch_chains_page"))
    os.makedirs(version_dir, exist_ok=True)
    # Always save as upgrade.tar.gz regardless of the uploaded file's original name
    file.save(os.path.join(version_dir, core.UPGRADE_FILENAME))

    flash(f"Added version {version} for {ctype}.", "success")
    return redirect(url_for("patch_chains_page"))


@app.route("/patch-chains/delete_step", methods=["POST"])
@admin_required
def patch_chains_delete_step():
    ctype = request.form.get("controller_type", "")
    version = request.form.get("version", "")

    # SECURITY: previously ctype/version were dropped straight into
    # os.path.join(...) and handed to shutil.rmtree() with NO validation —
    # a controller_type/version like "../../../../SomeOtherFolder" would
    # let an authenticated admin (or a hijacked admin session, e.g. via
    # CSRF) delete arbitrary directories anywhere the process can reach,
    # not just inside ControllerPatches/. Lock ctype to a known type and
    # collapse version to a bare folder name (no slashes, no "..").
    chains = core.load_patch_chains()
    if ctype not in chains:
        flash("Unknown controller type.", "error")
        return redirect(url_for("patch_chains_page"))

    version = os.path.basename((version or "").strip())
    if not version or version in (".", ".."):
        flash("Invalid version.", "error")
        return redirect(url_for("patch_chains_page"))

    base = os.path.realpath(core.CONTROLLER_PATCHES_DIR)
    version_dir = os.path.realpath(os.path.join(base, ctype.replace(" ", "_"), version))
    if not (version_dir == base or version_dir.startswith(base + os.sep)):
        flash("Invalid version path.", "error")
        return redirect(url_for("patch_chains_page"))

    if os.path.isdir(version_dir):
        import shutil as _shutil
        _shutil.rmtree(version_dir, ignore_errors=True)
        flash(f"Removed version {version}.", "success")
    else:
        flash("Could not find that version folder.", "error")
    return redirect(url_for("patch_chains_page"))


# ==============================================================================
# CFU (admin only) — settings for Job 6 (CFU APK deploy) + APK file upload
# ==============================================================================

@app.route("/cfu", methods=["GET", "POST"])
@admin_required
def cfu_page():
    if request.method == "POST":
        s = core.load_cfu_settings()
        s["remote_apk_dir"] = request.form.get("remote_apk_dir", "/sdcard/Download/").strip()
        s["config_path"] = request.form.get("config_path", "/sdcard/cfu/configuration.txt").strip()
        s["wifi_password"] = request.form.get("wifi_password", "").strip()
        s["dns1_default"] = request.form.get("dns1_default", "1.1.1.1").strip()
        s["dns2_default"] = request.form.get("dns2_default", "8.8.4.4").strip()
        s["app_component"] = request.form.get("app_component", "").strip()
        s["launch_app_after_install"] = request.form.get("launch_app_after_install") == "on"
        core.save_cfu_settings(s)
        flash("CFU settings saved.", "success")
        return redirect(url_for("cfu_page"))

    s = core.load_cfu_settings()
    apk_status = local_file_status(core.get_local_apk_path(s))
    return render_template("cfu.html", s=s, apk_status=apk_status)


@app.route("/cfu/upload_apk", methods=["POST"])
@admin_required
def cfu_upload_apk():
    file = request.files.get("apk_file")
    if not file or not file.filename:
        flash("No APK file selected.", "error")
        return redirect(url_for("cfu_page"))
    filename = secure_filename(file.filename)
    if not filename or filename.lower() in _PROTECTED_FILENAMES:
        flash("That file name isn't valid.", "error")
        return redirect(url_for("cfu_page"))
    dest = os.path.join(core.BASE_DIR, filename)
    file.save(dest)
    s = core.load_cfu_settings()
    s["apk_filename"] = filename
    core.save_cfu_settings(s)
    flash(f"Uploaded '{filename}' ({os.path.getsize(dest):,} bytes) and set it as the CFU APK.", "success")
    return redirect(url_for("cfu_page"))


# ==============================================================================
# Built-in Scheduler (admin only) — triggers runs automatically without
# needing an external Windows Task Scheduler entry.
# ==============================================================================

DAY_NAMES = ["mon", "tue", "wed", "thu", "fri", "sat", "sun"]

@app.route("/scheduler", methods=["GET", "POST"])
@admin_required
def scheduler_page():
    if request.method == "POST":
        try:
            hour12 = int(request.form.get("hour", "2"))
            minute = int(request.form.get("minute", "0"))
            ampm = request.form.get("ampm", "AM").upper()
            if not (1 <= hour12 <= 12 and 0 <= minute <= 59 and ampm in ("AM", "PM")):
                raise ValueError
        except ValueError:
            flash("That time doesn't look valid — pick an hour, minute, and AM/PM.", "error")
            return redirect(url_for("scheduler_page"))

        hour24 = hour12 % 12  # 12 AM -> 0, 12 PM -> 12
        if ampm == "PM":
            hour24 += 12
        normalized_time = f"{hour24:02d}:{minute:02d}"
        display_time = f"{hour12}:{minute:02d} {ampm}"

        mode = request.form.get("mode", "recurring")
        if mode not in ("recurring", "once"):
            mode = "recurring"

        run_date = request.form.get("run_date", "").strip()
        if mode == "once":
            try:
                time.strptime(run_date, "%Y-%m-%d")
            except ValueError:
                flash("Pick a valid date for a one-time schedule.", "error")
                return redirect(url_for("scheduler_page"))

        sched = {
            "enabled": request.form.get("enabled") == "on",
            "mode": mode,
            "time": normalized_time,
            "days": request.form.getlist("days") or DAY_NAMES,
            "run_date": run_date,
            "run_ppt": request.form.get("run_ppt") == "on",
            "run_shot": request.form.get("run_shot") == "on",
            "run_fr": request.form.get("run_fr") == "on",
            "run_patch": request.form.get("run_patch") == "on",
            "run_chain": request.form.get("run_chain") == "on",
            "resume": request.form.get("resume") == "on",
            "device_types": request.form.getlist("device_types"),
            "site_filter": request.form.getlist("site_filter"),
        }
        existing = core.load_schedule()
        # Reset the fired-guard whenever the target date/time/mode actually changes,
        # so editing a one-time schedule (e.g. after it already fired) re-arms it.
        same_target = (existing.get("mode") == mode and existing.get("time") == normalized_time
                       and existing.get("run_date") == run_date and existing.get("days") == sched["days"])
        sched["last_run_date"] = existing.get("last_run_date", "") if same_target else ""
        core.save_schedule(sched)
        if mode == "once":
            flash(f"Schedule saved — will trigger once on {run_date} at {display_time} (server local time).", "success")
        else:
            flash(f"Schedule saved — will trigger at {display_time} (server local time).", "success")
        return redirect(url_for("scheduler_page"))

    sched = core.load_schedule()
    # Convert stored 24h "HH:MM" into 12h hour/minute/AM-PM for the dropdowns
    hour24, minute = 2, 0
    parsed = core.parse_hm_to_minutes(sched.get("time", "02:00"))
    if parsed is not None:
        hour24, minute = parsed // 60, parsed % 60
    ampm = "AM" if hour24 < 12 else "PM"
    hour12 = hour24 % 12
    if hour12 == 0:
        hour12 = 12

    next_run_text = _describe_next_run(sched)
    server_now = time.strftime("%A, %b %d, %Y — %I:%M %p")
    today_date = time.strftime("%Y-%m-%d")

    return render_template("scheduler.html", sched=sched, day_names=DAY_NAMES,
                           device_types=core.DEVICE_TYPES, sites=core.distinct_sites(),
                           sel_hour=hour12, sel_minute=minute, sel_ampm=ampm,
                           next_run_text=next_run_text, server_now=server_now, today_date=today_date)


def _describe_next_run(sched):
    """Human-readable description of when the schedule will next fire, or why it won't."""
    if not sched.get("enabled"):
        return "Disabled — turn on \"Enable scheduled runs\" to activate."

    target_minutes = core.parse_hm_to_minutes(sched.get("time"))
    if target_minutes is None:
        return "Not fully configured yet — set a valid time."
    hh, mm = target_minutes // 60, target_minutes % 60
    ampm = "AM" if hh < 12 else "PM"
    h12 = hh % 12 or 12
    time_label = f"{h12}:{mm:02d} {ampm}"

    now = time.localtime()
    now_minutes = now.tm_hour * 60 + now.tm_min
    today_str = time.strftime("%Y-%m-%d", now)

    if sched.get("mode") == "once":
        run_date = sched.get("run_date", "")
        if not run_date:
            return "Not fully configured yet — pick a date."
        if sched.get("last_run_date"):
            return f"Already triggered on {sched['last_run_date']} — set a new date to schedule again."
        if run_date < today_str or (run_date == today_str and now_minutes >= target_minutes):
            return f"{run_date} at {time_label} has already passed and never fired — check the date/time."
        try:
            date_struct = time.strptime(run_date, "%Y-%m-%d")
            date_label = time.strftime("%A, %b %d, %Y", date_struct)
        except ValueError:
            date_label = run_date
        when = "Today" if run_date == today_str else date_label
        return f"{when} at {time_label} (one-time)"

    # Recurring mode
    days = sched.get("days") or []
    if not days:
        return "Not fully configured yet — pick at least one day."
    already_ran_today = sched.get("last_run_date") == today_str

    for offset in range(8):
        check_time = time.localtime(time.mktime(now) + offset * 86400)
        check_day = DAY_NAMES[check_time.tm_wday]
        if check_day not in days:
            continue
        if offset == 0:
            if already_ran_today or now_minutes >= target_minutes:
                continue  # today's slot is used up or already passed
        date_label = time.strftime("%A, %b %d", check_time)
        when = "Today" if offset == 0 else ("Tomorrow" if offset == 1 else date_label)
        return f"{when} at {time_label}"

    return "No matching day found in the next week — check your day selection."


def _scheduler_loop():
    """
    Background thread: checks every 30s whether the configured time has
    been reached yet. Uses "has the target time passed?" rather than an
    exact-minute string match — exact matching is fragile (a slow tick,
    a busy moment, or a time value like '2:30' instead of '02:30' could
    make it miss the single 60-second window it was checking for).

    Two modes:
      recurring — fires on any checked day-of-week, once per calendar day
                  (guarded by last_run_date == today).
      once      — fires a single time on a specific date, then never
                  again until the date/time/mode is changed (guarded by
                  last_run_date being non-empty at all).
    """
    while True:
        try:
            sched = core.load_schedule()
            if sched.get("enabled"):
                now = time.localtime()
                now_minutes = now.tm_hour * 60 + now.tm_min
                target_minutes = core.parse_hm_to_minutes(sched.get("time"))
                today_str = time.strftime("%Y-%m-%d", now)
                mode = sched.get("mode", "recurring")

                if target_minutes is None:
                    print(f"[Scheduler] Could not parse configured time: {sched.get('time')!r}")
                elif mode == "once":
                    due = (sched.get("run_date") == today_str and now_minutes >= target_minutes
                          and not sched.get("last_run_date") and not RUN_STATE.running)
                    if due:
                        sched["last_run_date"] = today_str
                        core.save_schedule(sched)
                        RUN_STATE.log(f"[Scheduler] Auto-starting one-time scheduled run "
                                     f"(target {sched.get('run_date')} {sched.get('time')})")
                        RUN_STATE.start(
                            sched.get("run_ppt", True), sched.get("run_shot", True),
                            sched.get("run_fr", True), sched.get("run_patch", False),
                            sched.get("run_chain", False), sched.get("resume", True),
                            sched.get("device_types") or list(core.DEVICE_TYPES),
                            True, "scheduler",
                            site_filter=sched.get("site_filter") or None,
                        )
                else:
                    today_day = DAY_NAMES[now.tm_wday]
                    already_ran_today = sched.get("last_run_date") == today_str
                    if (now_minutes >= target_minutes and today_day in sched.get("days", [])
                            and not already_ran_today and not RUN_STATE.running):
                        sched["last_run_date"] = today_str
                        core.save_schedule(sched)
                        RUN_STATE.log(f"[Scheduler] Auto-starting scheduled run (target {sched.get('time')}, now {time.strftime('%H:%M', now)})")
                        RUN_STATE.start(
                            sched.get("run_ppt", True), sched.get("run_shot", True),
                            sched.get("run_fr", True), sched.get("run_patch", False),
                            sched.get("run_chain", False), sched.get("resume", True),
                            sched.get("device_types") or list(core.DEVICE_TYPES),
                            True, "scheduler",
                            site_filter=sched.get("site_filter") or None,
                        )
        except Exception as e:
            print(f"[Scheduler error] {e}")
        time.sleep(30)


threading.Thread(target=_scheduler_loop, daemon=True).start()


# ==============================================================================
# Config backup / restore (admin only)
# ==============================================================================

@app.route("/settings/export_config")
@admin_required
def export_config():
    dest = os.path.join(core.BASE_DIR, "config_backup.zip")
    core.export_config_bundle(dest)
    return send_file(dest, as_attachment=True,
                     download_name=f"kiosk_config_backup_{time.strftime('%Y%m%d_%H%M%S')}.zip")


@app.route("/settings/import_config", methods=["POST"])
@admin_required
def import_config():
    file = request.files.get("config_zip")
    if not file or not file.filename:
        flash("No backup file selected.", "error")
        return redirect(url_for("settings_page"))
    tmp_path = os.path.join(core.BASE_DIR, "_import_tmp.zip")
    file.save(tmp_path)
    try:
        restored = core.import_config_bundle(tmp_path)
        flash(f"Restored: {', '.join(restored) if restored else '(nothing matched)'}", "success")
    except Exception as e:
        flash(f"Restore failed: {e}", "error")
    finally:
        if os.path.exists(tmp_path):
            os.remove(tmp_path)
    return redirect(url_for("settings_page"))


if __name__ == "__main__":
    from waitress import serve
    print("Wavetec Kiosk Web Dashboard starting on http://0.0.0.0:5000")
    print("Default login: admin / admin  (you'll be asked to change it on first login)")
    serve(app, host="0.0.0.0", port=5000, threads=8)
